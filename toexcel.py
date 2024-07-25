import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from tkinter import Tk, filedialog, Label
from tkinter.ttk import Progressbar
import controller as db_controller

def save_data_to_excel(data, file_name, columns, column_widths, progress_callback):
    try:
        # Convert the data to a pandas DataFrame
        df = pd.DataFrame(data, columns=columns)
        
        # Write the DataFrame to an Excel file
        df.to_excel(file_name, index=False, engine='openpyxl')
        
        # Load the workbook and select the active worksheet
        workbook = load_workbook(file_name)
        worksheet = workbook.active
        
        # Add colors and styles
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow header
        header_font = Font(bold=True)  # Bold font
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Set column widths
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Save the workbook
        workbook.save(file_name)
        
        # Call the progress callback
        progress_callback(100, "Saved successfully!")  # Update progress and message
        
        # Get the absolute path
        absolute_path = os.path.abspath(file_name)
        print(f"Absolute path to the file: {absolute_path}")
    except Exception as e:
        progress_callback(0, f"Error saving file: {e}")  # Update progress and message

def on_save_excel_button_click(data, columns, column_widths):
    file_name = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        title="Save file"
    )
    
    if file_name:
        # Start a new Tkinter window to show progress
        progress_window = Tk()
        progress_window.title("Saving File")
        
        # Create and configure the progress bar
        progress = Progressbar(progress_window, orient="horizontal", length=300, mode="determinate")
        progress.pack(pady=20)
        
        # Create and configure the message label
        message_label = Label(progress_window, text="Processing...", font=("Helvetica", 12))
        message_label.pack(pady=10)
        
        # Center the progress bar and message label
        progress_window.update_idletasks()
        window_width = progress_window.winfo_reqwidth()
        window_height = progress_window.winfo_reqheight()
        screen_width = progress_window.winfo_screenwidth()
        screen_height = progress_window.winfo_screenheight()
        x = (screen_width // 2) - (window_width // 2)
        y = (screen_height // 2) - (window_height // 2)
        progress_window.geometry(f"+{x}+{y}")
        
        def update_progress(value, message="Processing..."):
            progress['value'] = value
            message_label.config(text=message)
            progress_window.update_idletasks()
        
        # Call save_data_to_excel and update the progress
        save_data_to_excel(data, file_name, columns, column_widths, update_progress)
        
        # Schedule the window to close after 2 seconds to allow the user to read the message
        progress_window.after(2000, progress_window.destroy)  # Close after 2 seconds
